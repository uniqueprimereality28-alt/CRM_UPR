from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import csv
import io
import math
import asyncio
import logging
import uuid
from datetime import datetime, timezone, timedelta, date as _date
from typing import List, Optional, Annotated, Any

import bcrypt
import jwt
from bson import ObjectId
from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, BeforeValidator, ConfigDict
from starlette.middleware.cors import CORSMiddleware
from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

# --- NEW IMPORT for PDF report generation ---
from reports_pdf import generate_daily_report_pdf

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("crm")

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_ALGORITHM = "HS256"
LEAD_STATUSES = ["new", "contacted", "qualified", "site_visit", "negotiation", "won", "lost"]
FOLLOWUP_STATUSES = ["interested", "visit_scheduled", "not_interested", "callback", "converted"]
LEAD_TAGS = ["hot", "raw", "warm", "cold"]  # base tags; custom strings allowed too

# Roles hierarchy
ROLE_SUPER = "superadmin"
ROLE_ADMIN = "admin"
ROLE_HR = "hr"
ROLE_TL = "team_lead"
ROLE_SALES = "sales"
ROLE_EMP = "employee"
ALL_ROLES = {ROLE_SUPER, ROLE_ADMIN, ROLE_HR, ROLE_TL, ROLE_SALES, ROLE_EMP}
MANAGER_ROLES = {ROLE_SUPER, ROLE_ADMIN, ROLE_TL}  # can manage leads, view team
FULL_VIEW_ROLES = {ROLE_SUPER, ROLE_ADMIN, ROLE_HR}
ATTENDANCE_REQUIRED_ROLES = {ROLE_SUPER, ROLE_TL, ROLE_SALES, ROLE_EMP, ROLE_HR}

DEFAULT_PWD = os.environ.get("DEFAULT_USER_PASSWORD", "Welcome@123")

app = FastAPI(title="Unique Prime Reality CRM")
api = APIRouter(prefix="/api")


# ---------------- Mongo helpers ----------------
def _to_str(v: Any) -> Any:
    return str(v) if isinstance(v, ObjectId) else v


PyObjectId = Annotated[str, BeforeValidator(_to_str)]


class BaseDocument(BaseModel):
    model_config = ConfigDict(populate_by_name=True, extra="ignore")
    id: Optional[PyObjectId] = Field(default=None, alias="_id", serialization_alias="id")

    def to_mongo(self) -> dict:
        return self.model_dump(exclude_none=True, exclude={"id"})

    @classmethod
    def from_mongo(cls, doc: Optional[dict]):
        if not doc:
            return None
        return cls.model_validate(doc)


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def haversine_m(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    R = 6371000.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lng2 - lng1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))


def ist_now() -> datetime:
    return datetime.now(timezone.utc) + timedelta(minutes=330)


def ist_today_str() -> str:
    return ist_now().strftime("%Y-%m-%d")


# ---------------- Models ----------------
class UserPublic(BaseDocument):
    username: str
    name: str
    role: str
    email: Optional[str] = None
    phone: Optional[str] = None
    team_lead_id: Optional[str] = None
    team_lead_name: Optional[str] = None
    office_start: Optional[str] = None
    office_end: Optional[str] = None
    working_days: Optional[List[int]] = None  # 0=Mon..6=Sun
    avatar_url: Optional[str] = None
    date_of_birth: Optional[str] = None
    joining_date: Optional[str] = None
    joining_date_edits: int = 0
    attendance_exempt: bool = False
    wfh: bool = False
    active: bool = True
    created_at: Optional[str] = None


class LoginRequest(BaseModel):
    username: str
    password: str


class UserCreate(BaseModel):
    username: str
    name: str
    password: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    role: str = ROLE_SALES
    team_lead_id: Optional[str] = None
    office_start: Optional[str] = None
    office_end: Optional[str] = None
    working_days: Optional[List[int]] = None
    date_of_birth: Optional[str] = None
    joining_date: Optional[str] = None
    wfh: Optional[bool] = None
    attendance_exempt: Optional[bool] = None


class UserUpdate(BaseModel):
    name: Optional[str] = None
    password: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    role: Optional[str] = None
    team_lead_id: Optional[str] = None
    office_start: Optional[str] = None
    office_end: Optional[str] = None
    working_days: Optional[List[int]] = None
    active: Optional[bool] = None
    avatar_url: Optional[str] = None
    date_of_birth: Optional[str] = None
    joining_date: Optional[str] = None
    wfh: Optional[bool] = None
    attendance_exempt: Optional[bool] = None


class ProfileUpdate(BaseModel):
    """Self-service edits users can make to their own profile."""
    name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    avatar_url: Optional[str] = None
    date_of_birth: Optional[str] = None
    joining_date: Optional[str] = None


class PasswordChange(BaseModel):
    current_password: str
    new_password: str


class Lead(BaseDocument):
    name: str
    phone: str
    email: Optional[str] = None
    source: str = "Website"
    status: str = "new"
    tag: Optional[str] = None
    follow_up_status: Optional[str] = None
    budget: Optional[float] = None
    property_interest: Optional[str] = None
    city: Optional[str] = None
    notes: Optional[str] = None
    remark: Optional[str] = None  # custom per-lead tag/remark
    assigned_to: Optional[str] = None
    assigned_to_name: Optional[str] = None
    created_at: Optional[str] = None
    updated_at: Optional[str] = None
    last_contacted_at: Optional[str] = None
    total_talk_time: int = 0
    call_count: int = 0
    follow_up_at: Optional[str] = None
    follow_up_note: Optional[str] = None
    reminder_acked: bool = False
    brochure_sent: bool = False
    brochure_sent_at: Optional[str] = None


class LeadCreate(BaseModel):
    name: str
    phone: str
    email: Optional[str] = None
    source: str = "Website"
    status: str = "new"
    tag: Optional[str] = None
    follow_up_status: Optional[str] = None
    budget: Optional[float] = None
    property_interest: Optional[str] = None
    city: Optional[str] = None
    notes: Optional[str] = None
    remark: Optional[str] = None
    assigned_to: Optional[str] = None


class LeadUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    source: Optional[str] = None
    status: Optional[str] = None
    tag: Optional[str] = None
    follow_up_status: Optional[str] = None
    budget: Optional[float] = None
    property_interest: Optional[str] = None
    city: Optional[str] = None
    notes: Optional[str] = None
    remark: Optional[str] = None
    assigned_to: Optional[str] = None
    follow_up_at: Optional[str] = None
    follow_up_note: Optional[str] = None
    brochure_sent: Optional[bool] = None


class AssignRequest(BaseModel):
    lead_ids: List[str]
    agent_id: str


class ActivityCreate(BaseModel):
    type: str = "note"
    message: str


class CallStart(BaseModel):
    lead_id: str


class CallEnd(BaseModel):
    duration: int
    outcome: str = "connected"
    notes: Optional[str] = None
    follow_up_at: Optional[str] = None
    follow_up_note: Optional[str] = None


class AttendanceCheckIn(BaseModel):
    lat: float
    lng: float
    accuracy: Optional[float] = None


class AttendanceCheckOut(BaseModel):
    lat: float
    lng: float
    accuracy: Optional[float] = None


class AttendanceEdit(BaseModel):
    check_in_at: Optional[str] = None
    check_out_at: Optional[str] = None
    note: Optional[str] = None
    status: Optional[str] = None  # "present" | "absent" | "leave"


class SettingsUpdate(BaseModel):
    office_lat: Optional[float] = None
    office_lng: Optional[float] = None
    office_radius_m: Optional[int] = None
    office_start: Optional[str] = None
    office_end: Optional[str] = None
    office_label: Optional[str] = None


class SystemResetRequest(BaseModel):
    confirm: str


class GroupCreate(BaseModel):
    name: str
    member_ids: List[str] = []


class GroupUpdate(BaseModel):
    name: Optional[str] = None
    member_ids: Optional[List[str]] = None


class MessageCreate(BaseModel):
    text: str


class AlertCreate(BaseModel):
    message: str
    target: str = "all"  # "all" | role | user_id
    priority: str = "normal"  # normal | high
    duration_hours: int = 24  # alert auto-expires after this many hours


# ---------------- Storage (local disk) ----------------
STORAGE_APP_PREFIX = "upr-crm"
STORAGE_ROOT = Path(__file__).parent / "storage"
STORAGE_ROOT.mkdir(parents=True, exist_ok=True)


def _put_object(path: str, data: bytes, content_type: str) -> dict:
    full_path = STORAGE_ROOT / path
    full_path.parent.mkdir(parents=True, exist_ok=True)
    full_path.write_bytes(data)
    (full_path.parent / f"{full_path.name}.meta").write_text(content_type)
    return {"path": path}


def _get_object(path: str) -> tuple:
    full_path = STORAGE_ROOT / path
    if not full_path.exists():
        raise FileNotFoundError(path)
    meta_path = full_path.parent / f"{full_path.name}.meta"
    content_type = meta_path.read_text() if meta_path.exists() else "application/octet-stream"
    return full_path.read_bytes(), content_type


# ---------------- Auth ----------------
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plain: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))
    except (ValueError, TypeError):
        return False


def create_access_token(user_id: str, username: str, role: str) -> str:
    payload = {
        "sub": user_id, "username": username, "role": role, "type": "access",
        "exp": datetime.now(timezone.utc) + timedelta(hours=12),
    }
    return jwt.encode(payload, os.environ["JWT_SECRET"], algorithm=JWT_ALGORITHM)


async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        header = request.headers.get("Authorization", "")
        if header.startswith("Bearer "):
            token = header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, os.environ["JWT_SECRET"], algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Session expired, please login again")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")
    user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
    if not user or not user.get("active", True):
        raise HTTPException(status_code=401, detail="User not found or deactivated")
    return user


def is_super(u: dict) -> bool: return u.get("role") == ROLE_SUPER
def is_admin_or_super(u: dict) -> bool: return u.get("role") in {ROLE_SUPER, ROLE_ADMIN}
def is_manager(u: dict) -> bool: return u.get("role") in MANAGER_ROLES
def can_view_all(u: dict) -> bool: return u.get("role") in FULL_VIEW_ROLES


async def require_admin(user: dict = Depends(get_current_user)) -> dict:
    if not is_admin_or_super(user):
        raise HTTPException(status_code=403, detail="Admin access required")
    return user


async def require_super(user: dict = Depends(get_current_user)) -> dict:
    # Sandeep (admin) and Vranda are equal top-level administrators; both may edit anything.
    if not is_admin_or_super(user):
        raise HTTPException(status_code=403, detail="Administrator access required")
    return user


async def require_superadmin_only(user: dict = Depends(get_current_user)) -> dict:
    # Strictly Vranda's role (superadmin) — used for the irreversible system data reset,
    # which is intentionally NOT available to the admin (Sandeep) account.
    if not is_super(user):
        raise HTTPException(status_code=403, detail="Superadmin access required")
    return user


async def require_manager(user: dict = Depends(get_current_user)) -> dict:
    if not is_manager(user):
        raise HTTPException(status_code=403, detail="Managerial access required")
    return user


async def _team_member_ids(user: dict) -> List[str]:
    """Return the set of user_ids visible to `user`."""
    if can_view_all(user):
        rows = await db.users.find({}).to_list(2000)
        return [str(r["_id"]) for r in rows]
    if user.get("role") == ROLE_TL:
        members = await db.users.find({"team_lead_id": str(user["_id"])}).to_list(500)
        return [str(user["_id"])] + [str(m["_id"]) for m in members]
    return [str(user["_id"])]


@api.post("/auth/login")
async def login(payload: LoginRequest, response: Response):
    username = payload.username.strip().lower()
    attempt = await db.login_attempts.find_one({"identifier": username})
    if attempt and attempt.get("count", 0) >= 5:
        locked_until = attempt.get("locked_until")
        if locked_until and datetime.fromisoformat(locked_until) > datetime.now(timezone.utc):
            raise HTTPException(status_code=429, detail="Too many failed attempts. Try again in 15 minutes.")

    user = await db.users.find_one({"username": username})
    if not user or not verify_password(payload.password, user["password_hash"]):
        await db.login_attempts.update_one(
            {"identifier": username},
            {"$inc": {"count": 1},
             "$set": {"locked_until": (datetime.now(timezone.utc) + timedelta(minutes=15)).isoformat()}},
            upsert=True)
        raise HTTPException(status_code=401, detail="Invalid username or password")
    if not user.get("active", True):
        raise HTTPException(status_code=403, detail="Account deactivated. Contact admin.")

    await db.login_attempts.delete_one({"identifier": username})
    token = create_access_token(str(user["_id"]), user["username"], user["role"])
    secure = os.environ.get("COOKIE_SECURE", "true").lower() == "true"
    response.set_cookie("access_token", token, httponly=True, secure=secure,
                        samesite="none" if secure else "lax", max_age=43200, path="/")
    await db.users.update_one({"_id": user["_id"]}, {"$set": {"last_login": now_iso()}})
    return {"access_token": token, "user": UserPublic.from_mongo(user)}


@api.get("/auth/me", response_model=UserPublic)
async def me(user: dict = Depends(get_current_user)):
    return UserPublic.from_mongo(user)


@api.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}


@api.post("/auth/change-password")
async def change_password(payload: PasswordChange, user: dict = Depends(get_current_user)):
    if not verify_password(payload.current_password, user["password_hash"]):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    if len(payload.new_password) < 6:
        raise HTTPException(status_code=400, detail="New password must be at least 6 characters")
    await db.users.update_one({"_id": user["_id"]},
                              {"$set": {"password_hash": hash_password(payload.new_password),
                                        "password_changed_at": now_iso()}})
    return {"ok": True}


@api.put("/auth/profile")
async def update_own_profile(payload: ProfileUpdate, user: dict = Depends(get_current_user)):
    updates = payload.model_dump(exclude_none=True)
    if not updates:
        return UserPublic.from_mongo(user)
    # Joining-date rule: user may edit their own joining_date only TWICE.
    # After that, only an administrator can change it via PUT /users/{id}.
    if "joining_date" in updates and updates["joining_date"] != user.get("joining_date"):
        edits = int(user.get("joining_date_edits", 0))
        if edits >= 2 and not is_admin_or_super(user):
            raise HTTPException(
                status_code=403,
                detail="You've already updated your joining date twice. Ask an administrator to fix it for you.",
            )
        updates["joining_date_edits"] = edits + 1
    await db.users.update_one({"_id": user["_id"]}, {"$set": updates})
    if updates.get("name"):
        await db.leads.update_many({"assigned_to": str(user["_id"])},
                                   {"$set": {"assigned_to_name": updates["name"]}})
    doc = await db.users.find_one({"_id": user["_id"]})
    return UserPublic.from_mongo(doc)


AVATAR_MAX_BYTES = 3 * 1024 * 1024
AVATAR_MIMES = {"image/jpeg": "jpg", "image/png": "png", "image/webp": "webp", "image/gif": "gif"}
AVATAR_DIR = STORAGE_ROOT / "avatars"
AVATAR_DIR.mkdir(parents=True, exist_ok=True)


@api.post("/auth/avatar")
async def upload_avatar(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    content_type = (file.content_type or "").split(";")[0].strip().lower()
    ext = AVATAR_MIMES.get(content_type)
    if not ext:
        raise HTTPException(status_code=400, detail=f"Unsupported image format: {content_type or 'unknown'}")
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Empty image file")
    if len(data) > AVATAR_MAX_BYTES:
        raise HTTPException(status_code=413, detail="Image too large (max 3 MB)")
    fname = f"{user['_id']}-{uuid.uuid4().hex[:8]}.{ext}"
    (AVATAR_DIR / fname).write_bytes(data)
    (AVATAR_DIR / f"{fname}.meta").write_text(content_type)
    url = f"/api/avatars/{fname}"
    await db.users.update_one({"_id": user["_id"]}, {"$set": {"avatar_url": url}})
    return {"avatar_url": url}


@api.get("/avatars/{fname}")
async def serve_avatar(fname: str):
    path = AVATAR_DIR / fname
    if not path.exists() or "/" in fname or ".." in fname:
        raise HTTPException(status_code=404, detail="Avatar not found")
    meta_path = AVATAR_DIR / f"{fname}.meta"
    ct = meta_path.read_text() if meta_path.exists() else "image/jpeg"
    return Response(content=path.read_bytes(), media_type=ct,
                    headers={"Cache-Control": "public, max-age=86400"})


# ---------------- Users ----------------
@api.get("/users", response_model=List[UserPublic])
async def list_users(user: dict = Depends(get_current_user), role: Optional[str] = None,
                     team_lead_id: Optional[str] = None):
    query: dict = {}
    if role:
        query["role"] = role
    if team_lead_id:
        query["team_lead_id"] = team_lead_id
    # Scope for team leads: they see themselves + their team
    if not can_view_all(user) and user.get("role") == ROLE_TL:
        ids = await _team_member_ids(user)
        query["_id"] = {"$in": [ObjectId(i) for i in ids]}
    elif not can_view_all(user):
        query["_id"] = user["_id"]
    docs = await db.users.find(query).sort("created_at", -1).to_list(2000)
    return [UserPublic.from_mongo(d) for d in docs]


def _validate_role_creation(creator_role: str, target_role: str):
    if target_role not in ALL_ROLES:
        raise HTTPException(status_code=400, detail=f"Invalid role: {target_role}")
    if creator_role in {ROLE_SUPER, ROLE_ADMIN}:
        return  # top-level admins can create any role, including peer administrators
    if creator_role == ROLE_TL:
        if target_role != ROLE_SALES:
            raise HTTPException(status_code=403, detail="Team leaders can only create sales accounts")
        return
    raise HTTPException(status_code=403, detail="You cannot create users")


@api.post("/users", response_model=UserPublic)
async def create_user(payload: UserCreate, actor: dict = Depends(get_current_user)):
    _validate_role_creation(actor.get("role"), payload.role)
    username = payload.username.strip().lower()
    if await db.users.find_one({"username": username}):
        raise HTTPException(status_code=400, detail="Username already exists")

    team_lead_id = payload.team_lead_id
    team_lead_name = None
    # Team leaders auto-attach to themselves when creating sales
    if actor.get("role") == ROLE_TL and payload.role == ROLE_SALES:
        team_lead_id = str(actor["_id"])
    if team_lead_id:
        tl = await db.users.find_one({"_id": ObjectId(team_lead_id)})
        if not tl or tl.get("role") != ROLE_TL:
            raise HTTPException(status_code=400, detail="Assigned team lead is not valid")
        team_lead_name = tl.get("name")

    doc = {
        "username": username,
        "name": payload.name,
        "email": payload.email,
        "phone": payload.phone,
        "role": payload.role,
        "team_lead_id": team_lead_id,
        "team_lead_name": team_lead_name,
        "office_start": payload.office_start or os.environ.get("OFFICE_START", "11:00"),
        "office_end": payload.office_end or os.environ.get("OFFICE_END", "18:00"),
        "working_days": payload.working_days or [0, 1, 2, 3, 4, 5],
        "date_of_birth": payload.date_of_birth,
        "joining_date": payload.joining_date or ist_today_str(),
        "wfh": bool(payload.wfh),
        "attendance_exempt": bool(payload.attendance_exempt),
        "active": True,
        "password_hash": hash_password(payload.password or DEFAULT_PWD),
        "created_at": now_iso(),
        "created_by": str(actor["_id"]),
    }
    res = await db.users.insert_one(doc)
    doc["_id"] = res.inserted_id
    return UserPublic.from_mongo(doc)


@api.put("/users/{user_id}", response_model=UserPublic)
async def update_user(user_id: str, payload: UserUpdate, actor: dict = Depends(get_current_user)):
    target = await db.users.find_one({"_id": ObjectId(user_id)})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")

    # Permissions: admins (admin/super) can edit anyone including each other;
    # team_lead can edit their team members' basic profile info.
    if not is_admin_or_super(actor):
        if actor.get("role") == ROLE_TL and target.get("team_lead_id") == str(actor["_id"]):
            allowed = {"name", "email", "phone", "password", "avatar_url"}
            if any(k not in allowed for k, v in payload.model_dump(exclude_none=True).items()):
                raise HTTPException(status_code=403, detail="Team leads can only edit basic profile info")
        elif str(target["_id"]) == str(actor["_id"]):
            # anyone editing themselves is fine (limited to profile fields via /auth/profile normally)
            allowed = {"name", "email", "phone", "password", "avatar_url"}
            if any(k not in allowed for k, v in payload.model_dump(exclude_none=True).items()):
                raise HTTPException(status_code=403, detail="You can only edit basic profile info yourself")
        else:
            raise HTTPException(status_code=403, detail="You cannot edit this user")

    updates = {k: v for k, v in payload.model_dump(exclude_none=True).items() if k != "password"}
    if payload.password:
        updates["password_hash"] = hash_password(payload.password)
    # Only top-level admins can change role or team_lead_id assignments
    if not is_admin_or_super(actor):
        for k in ("role", "team_lead_id", "attendance_exempt", "wfh"):
            updates.pop(k, None)
    if "team_lead_id" in updates and updates["team_lead_id"]:
        tl = await db.users.find_one({"_id": ObjectId(updates["team_lead_id"])})
        if not tl or tl.get("role") != ROLE_TL:
            raise HTTPException(status_code=400, detail="Assigned team lead is not valid")
        updates["team_lead_name"] = tl.get("name")

    if updates:
        await db.users.update_one({"_id": ObjectId(user_id)}, {"$set": updates})
    if payload.name:
        await db.leads.update_many({"assigned_to": user_id}, {"$set": {"assigned_to_name": payload.name}})
    doc = await db.users.find_one({"_id": ObjectId(user_id)})
    return UserPublic.from_mongo(doc)


@api.delete("/users/{user_id}")
async def delete_user(user_id: str, actor: dict = Depends(require_admin)):
    target = await db.users.find_one({"_id": ObjectId(user_id)})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    if target.get("role") in {ROLE_SUPER, ROLE_ADMIN} and not is_admin_or_super(actor):
        raise HTTPException(status_code=403, detail="Only administrators can delete administrator accounts")
    await db.users.delete_one({"_id": ObjectId(user_id)})
    await db.leads.update_many({"assigned_to": user_id},
                               {"$set": {"assigned_to": None, "assigned_to_name": None}})
    return {"ok": True}


# ---------------- Activity helper ----------------
async def _log_activity(lead_id: str, actor: dict, type_: str, message: str, meta: Optional[dict] = None):
    await db.activities.insert_one({
        "lead_id": lead_id, "type": type_, "message": message,
        "actor_id": str(actor["_id"]), "actor_name": actor.get("name", "System"),
        "meta": meta or {}, "created_at": now_iso(),
    })


# ---------------- Leads ----------------
def _lead_visibility_query(user: dict) -> dict:
    """MongoDB query that restricts leads to what `user` may see."""
    if can_view_all(user):
        return {}
    if user.get("role") == ROLE_TL:
        # Own leads + leads of team members
        member_ids = [str(user["_id"])]
        # cache team members - callers should await; here return a $expr-like using assigned_to check later
        return {"__tl__": True}
    return {"assigned_to": str(user["_id"])}


async def _apply_visibility(query: dict, user: dict) -> dict:
    q = dict(query)
    if q.pop("__tl__", False):
        members = await db.users.find({"team_lead_id": str(user["_id"])}).to_list(500)
        ids = [str(user["_id"])] + [str(m["_id"]) for m in members]
        q["assigned_to"] = {"$in": ids}
    return q


@api.get("/leads", response_model=List[Lead])
async def list_leads(
    user: dict = Depends(get_current_user),
    status: Optional[str] = None,
    tag: Optional[str] = None,
    follow_up_status: Optional[str] = None,
    assigned_to: Optional[str] = None,
    source: Optional[str] = None,
    search: Optional[str] = None,
    unassigned: Optional[bool] = None,
    follow_up_from: Optional[str] = None,
    follow_up_to: Optional[str] = None,
    limit: int = 500,
):
    base_q = _lead_visibility_query(user)
    query = await _apply_visibility(base_q, user)

    if can_view_all(user) or user.get("role") == ROLE_TL:
        if assigned_to:
            query["assigned_to"] = assigned_to
        if unassigned:
            query["assigned_to"] = None
    if status:
        query["status"] = status
    if tag:
        query["tag"] = tag
    if follow_up_status:
        query["follow_up_status"] = follow_up_status
    if source:
        query["source"] = source
    if follow_up_from or follow_up_to:
        rng: dict = {}
        if follow_up_from:
            rng["$gte"] = follow_up_from
        if follow_up_to:
            rng["$lte"] = follow_up_to
        query["follow_up_at"] = rng
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
            {"city": {"$regex": search, "$options": "i"}},
            {"remark": {"$regex": search, "$options": "i"}},
        ]
    docs = await db.leads.find(query).sort("created_at", -1).to_list(limit)
    return [Lead.from_mongo(d) for d in docs]


@api.post("/leads", response_model=Lead)
async def create_lead(payload: LeadCreate, user: dict = Depends(get_current_user)):
    data = payload.model_dump()
    if data.get("status") not in LEAD_STATUSES:
        data["status"] = "new"
    assigned_to = data.get("assigned_to")
    if not assigned_to and user["role"] in {ROLE_SALES, ROLE_TL}:
        assigned_to = str(user["_id"])
    assigned_name = None
    if assigned_to:
        agent = await db.users.find_one({"_id": ObjectId(assigned_to)})
        assigned_name = agent.get("name") if agent else None
    lead = Lead(**{**data, "assigned_to": assigned_to, "assigned_to_name": assigned_name,
                   "created_at": now_iso(), "updated_at": now_iso()})
    doc = lead.to_mongo()
    res = await db.leads.insert_one(doc)
    doc["_id"] = res.inserted_id
    await _log_activity(str(res.inserted_id), user, "created", f"Lead created by {user.get('name')}")
    return Lead.from_mongo(doc)


async def _ensure_lead_access(lead: dict, user: dict):
    if can_view_all(user):
        return
    aid = lead.get("assigned_to")
    if user.get("role") == ROLE_TL:
        if aid == str(user["_id"]):
            return
        # allow if assigned to a team member
        if aid:
            member = await db.users.find_one({"_id": ObjectId(aid)})
            if member and member.get("team_lead_id") == str(user["_id"]):
                return
        raise HTTPException(status_code=403, detail="This lead isn't in your team")
    if aid != str(user["_id"]):
        raise HTTPException(status_code=403, detail="This lead is not assigned to you")


@api.get("/leads/{lead_id}")
async def get_lead(lead_id: str, user: dict = Depends(get_current_user)):
    doc = await db.leads.find_one({"_id": ObjectId(lead_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Lead not found")
    await _ensure_lead_access(doc, user)
    activities = await db.activities.find({"lead_id": lead_id}).sort("created_at", -1).to_list(300)
    calls = await db.calls.find({"lead_id": lead_id}).sort("started_at", -1).to_list(300)
    for a in activities: a["_id"] = str(a["_id"])
    for c in calls: c["_id"] = str(c["_id"])
    return {"lead": Lead.from_mongo(doc), "activities": activities, "calls": calls}


@api.put("/leads/{lead_id}", response_model=Lead)
async def update_lead(lead_id: str, payload: LeadUpdate, user: dict = Depends(get_current_user)):
    doc = await db.leads.find_one({"_id": ObjectId(lead_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Lead not found")
    await _ensure_lead_access(doc, user)
    updates = payload.model_dump(exclude_unset=True)
    if "status" in updates and updates["status"] not in LEAD_STATUSES:
        raise HTTPException(status_code=400, detail="Invalid status")
    if "assigned_to" in updates:
        if not is_manager(user):
            updates.pop("assigned_to")
        elif updates["assigned_to"]:
            agent = await db.users.find_one({"_id": ObjectId(updates["assigned_to"])})
            updates["assigned_to_name"] = agent.get("name") if agent else None
        else:
            updates["assigned_to_name"] = None
    if "brochure_sent" in updates:
        updates["brochure_sent_at"] = now_iso() if updates["brochure_sent"] else None
    if "follow_up_at" in updates:
        updates["reminder_acked"] = False
    updates["updated_at"] = now_iso()
    await db.leads.update_one({"_id": ObjectId(lead_id)}, {"$set": updates})
    if "status" in updates and updates["status"] != doc.get("status"):
        await _log_activity(lead_id, user, "status_change",
                            f"Status changed from {doc.get('status')} to {updates['status']}")
    if updates.get("assigned_to_name"):
        await _log_activity(lead_id, user, "assignment",
                            f"Lead assigned to {updates['assigned_to_name']}")
    if "tag" in updates:
        await _log_activity(lead_id, user, "tag", f"Tag set to {updates['tag']}")
    if "follow_up_status" in updates:
        await _log_activity(lead_id, user, "followup",
                            f"Follow-up status: {updates['follow_up_status']}")
    if "brochure_sent" in updates and updates["brochure_sent"] != doc.get("brochure_sent", False):
        await _log_activity(lead_id, user, "brochure",
                            "Brochure sent on WhatsApp" if updates["brochure_sent"] else "Brochure tick removed")
    if updates.get("follow_up_at") and updates["follow_up_at"] != doc.get("follow_up_at"):
        await _log_activity(lead_id, user, "followup", "Follow-up reminder scheduled")
    new_doc = await db.leads.find_one({"_id": ObjectId(lead_id)})
    return Lead.from_mongo(new_doc)


@api.delete("/leads/{lead_id}")
async def delete_lead(lead_id: str, admin: dict = Depends(require_admin)):
    await db.leads.delete_one({"_id": ObjectId(lead_id)})
    await db.activities.delete_many({"lead_id": lead_id})
    return {"ok": True}


@api.post("/leads/assign")
async def assign_leads(payload: AssignRequest, user: dict = Depends(require_manager)):
    agent = await db.users.find_one({"_id": ObjectId(payload.agent_id)})
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")
    # Team leads can only assign to their own team members (or themselves)
    if user.get("role") == ROLE_TL:
        if agent.get("team_lead_id") != str(user["_id"]) and str(agent["_id"]) != str(user["_id"]):
            raise HTTPException(status_code=403, detail="You can only assign leads to your own team")
    ids = [ObjectId(i) for i in payload.lead_ids]
    await db.leads.update_many(
        {"_id": {"$in": ids}},
        {"$set": {"assigned_to": payload.agent_id, "assigned_to_name": agent["name"],
                  "updated_at": now_iso()}})
    for lead_id in payload.lead_ids:
        await _log_activity(lead_id, user, "assignment", f"Lead assigned to {agent['name']}")
    return {"ok": True, "assigned": len(payload.lead_ids)}


@api.get("/leads/export")
async def export_leads(
    format: str = "xlsx",
    user: dict = Depends(require_manager),
    status: Optional[str] = None,
    tag: Optional[str] = None,
    assigned_to: Optional[str] = None,
):
    """Bulk export of visible leads to Excel or PDF. Tags/remarks columns included as-is, unchanged."""
    base_q = _lead_visibility_query(user)
    query = await _apply_visibility(base_q, user)
    if status:
        query["status"] = status
    if tag:
        query["tag"] = tag
    if assigned_to:
        query["assigned_to"] = assigned_to
    docs = await db.leads.find(query).sort("created_at", -1).to_list(20000)

    columns = ["Name", "Phone Number", "Email", "Status", "Tag", "Remark",
               "Source", "City", "Assigned To", "Created At"]

    def row_for(d):
        return [
            d.get("name") or "", d.get("phone") or "", d.get("email") or "",
            d.get("status") or "", d.get("tag") or "", d.get("remark") or "",
            d.get("source") or "", d.get("city") or "",
            d.get("assigned_to_name") or "Unassigned", (d.get("created_at") or "")[:10],
        ]

    if format == "pdf":
        buf = io.BytesIO()
        pdf_doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
        data = [columns] + [row_for(d) for d in docs]
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f4f6")]),
        ]))
        pdf_doc.build([table])
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=leads_{ist_today_str()}.pdf"})

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append(columns)
    for d in docs:
        ws.append(row_for(d))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=leads_{ist_today_str()}.xlsx"})


@api.get("/leads/import-template")
async def leads_import_template(user: dict = Depends(require_manager)):
    """Blank Excel template with just Name and Phone Number columns for clean imports."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append(["Name", "Phone Number"])
    ws.append(["Rahul Sharma", "9876543210"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=leads_import_template.xlsx"})


@api.post("/leads/import")
async def import_leads(
    file: UploadFile = File(...),
    assigned_to: Optional[str] = Form(None),
    default_tag: Optional[str] = Form(None),
    default_remark: Optional[str] = Form(None),
    skip_duplicates: Optional[str] = Form("true"),
    user: dict = Depends(require_manager),
):
    raw = await file.read()
    filename = (file.filename or "").lower()

    rows: List[dict] = []
    if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
        try:
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        except Exception:
            raise HTTPException(status_code=400,
                                detail="Could not read the Excel file. Please re-save it as .xlsx and try again.")
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = [str(h or "").strip().lower().replace(" ", "_") for h in next(rows_iter)]
        except StopIteration:
            header = []
        for r in rows_iter:
            row = {}
            for i, val in enumerate(r):
                if i < len(header) and header[i]:
                    row[header[i]] = "" if val is None else str(val).strip()
            if any(v for v in row.values()):
                rows.append(row)
    else:
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw.decode("latin-1")
        reader = csv.DictReader(io.StringIO(text))
        for r in reader:
            rows.append({(k or "").strip().lower().replace(" ", "_"): (v or "").strip() for k, v in r.items()})

    agent_name = None
    if assigned_to:
        agent = await db.users.find_one({"_id": ObjectId(assigned_to)})
        agent_name = agent.get("name") if agent else None

    skip_dupes = str(skip_duplicates).lower() in ("true", "1", "yes")
    existing = await db.leads.find({}, {"phone": 1, "name": 1}).to_list(50000)

    def norm_phone(p: str) -> str:
        return "".join(ch for ch in (p or "") if ch.isdigit())[-10:]

    existing_phones = {norm_phone(e.get("phone", "")): e.get("name") for e in existing if e.get("phone")}

    inserted, skipped, dupes, missing = 0, 0, [], 0
    seen_in_file = set()
    for row in rows:
        name = row.get("name") or row.get("full_name") or row.get("lead_name")
        phone = row.get("phone") or row.get("phone_number") or row.get("mobile") or row.get("contact")
        if not name or not phone:
            missing += 1
            continue
        np = norm_phone(phone)
        if np and (np in existing_phones or np in seen_in_file):
            dupes.append({"name": name, "phone": phone,
                          "existing_name": existing_phones.get(np) or "duplicated in the same file"})
            if skip_dupes:
                skipped += 1
                continue
        seen_in_file.add(np)
        budget = None
        try:
            if row.get("budget"):
                budget = float(str(row["budget"]).replace(",", ""))
        except ValueError:
            budget = None
        status = row.get("status", "new").lower()
        doc = Lead(
            name=name, phone=phone, email=row.get("email") or None,
            source=row.get("source") or "Excel Import",
            status=status if status in LEAD_STATUSES else "new",
            tag=row.get("tag") or default_tag or None,
            remark=row.get("remark") or default_remark or None,
            budget=budget,
            property_interest=row.get("property_interest") or row.get("property") or None,
            city=row.get("city") or None,
            notes=row.get("notes") or None,
            assigned_to=assigned_to, assigned_to_name=agent_name,
            created_at=now_iso(), updated_at=now_iso(),
        ).to_mongo()
        await db.leads.insert_one(doc)
        inserted += 1
    return {"inserted": inserted, "skipped": skipped, "missing": missing,
            "duplicates": dupes[:100], "total_duplicates": len(dupes)}


@api.post("/leads/{lead_id}/activities")
async def add_activity(lead_id: str, payload: ActivityCreate, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"_id": ObjectId(lead_id)})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    await _ensure_lead_access(lead, user)
    await _log_activity(lead_id, user, payload.type, payload.message)
    return {"ok": True}


# ---------------- Calls / recordings ----------------
RECORDING_MAX_BYTES = 30 * 1024 * 1024
RECORDING_MIMES = {
    "audio/webm": "webm", "audio/mp4": "m4a", "audio/mpeg": "mp3",
    "audio/ogg": "ogg", "audio/wav": "wav", "audio/x-m4a": "m4a",
    "audio/aac": "aac", "video/webm": "webm", "video/mp4": "m4a",
}


@api.get("/calls/config")
async def call_config(user: dict = Depends(get_current_user)):
    return {"mode": "manual", "recording_enabled": True}


@api.post("/calls/start")
async def start_call(payload: CallStart, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"_id": ObjectId(payload.lead_id)})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    await _ensure_lead_access(lead, user)
    doc = {
        "lead_id": payload.lead_id, "lead_name": lead["name"], "lead_phone": lead["phone"],
        "agent_id": str(user["_id"]), "agent_name": user.get("name"),
        "provider": "manual", "status": "in_progress", "duration": 0,
        "has_recording": False, "started_at": now_iso(),
    }
    res = await db.calls.insert_one(doc)
    await _log_activity(payload.lead_id, user, "call_started", f"Call initiated to {lead['phone']}")
    return {"call_id": str(res.inserted_id), "provider": "manual"}


@api.post("/calls/{call_id}/end")
async def end_call(call_id: str, payload: CallEnd, user: dict = Depends(get_current_user)):
    call = await db.calls.find_one({"_id": ObjectId(call_id)})
    if not call:
        raise HTTPException(status_code=404, detail="Call not found")
    duration = max(0, int(payload.duration))
    await db.calls.update_one(
        {"_id": ObjectId(call_id)},
        {"$set": {"status": "completed", "duration": duration, "outcome": payload.outcome,
                  "notes": payload.notes, "ended_at": now_iso()}})
    await db.leads.update_one(
        {"_id": ObjectId(call["lead_id"])},
        {"$inc": {"total_talk_time": duration, "call_count": 1},
         "$set": {"last_contacted_at": now_iso(), "updated_at": now_iso()}})
    mins, secs = duration // 60, duration % 60
    await _log_activity(call["lead_id"], user, "call_logged",
                        f"Call {payload.outcome} · {mins}m {secs}s" + (f" · {payload.notes}" if payload.notes else ""),
                        {"duration": duration})
    if payload.follow_up_at:
        await db.leads.update_one(
            {"_id": ObjectId(call["lead_id"])},
            {"$set": {"follow_up_at": payload.follow_up_at,
                      "follow_up_note": payload.follow_up_note,
                      "reminder_acked": False}})
        await _log_activity(call["lead_id"], user, "followup",
                            "Follow-up reminder scheduled after call"
                            + (f" · {payload.follow_up_note}" if payload.follow_up_note else ""))
    return {"ok": True, "duration": duration}


@api.get("/calls")
async def list_calls(user: dict = Depends(get_current_user), agent_id: Optional[str] = None, limit: int = 300):
    query: dict = {}
    if not can_view_all(user):
        if user.get("role") == ROLE_TL:
            ids = await _team_member_ids(user)
            query["agent_id"] = {"$in": ids}
        else:
            query["agent_id"] = str(user["_id"])
    elif agent_id:
        query["agent_id"] = agent_id
    docs = await db.calls.find(query).sort("started_at", -1).to_list(limit)
    for d in docs: d["_id"] = str(d["_id"])
    return docs


def _can_access_call(user: dict, call: dict) -> bool:
    if can_view_all(user):
        return True
    return call.get("agent_id") == str(user["_id"])


@api.delete("/calls/{call_id}")
async def delete_call(call_id: str, admin: dict = Depends(require_admin)):
    """Admins (admin/superadmin — e.g. Sandeep or Vranda) can permanently delete a call log."""
    call = await db.calls.find_one({"_id": ObjectId(call_id)})
    if not call:
        raise HTTPException(status_code=404, detail="Call not found")
    # Roll back any talk-time/call-count this completed call had added to the lead
    if call.get("status") == "completed" and call.get("lead_id"):
        await db.leads.update_one(
            {"_id": ObjectId(call["lead_id"])},
            {"$inc": {"total_talk_time": -int(call.get("duration") or 0), "call_count": -1}})
    await db.calls.delete_one({"_id": ObjectId(call_id)})
    return {"ok": True}


@api.post("/calls/{call_id}/recording")
async def upload_recording(call_id: str, file: UploadFile = File(...),
                           user: dict = Depends(get_current_user)):
    call = await db.calls.find_one({"_id": ObjectId(call_id)})
    if not call:
        raise HTTPException(status_code=404, detail="Call not found")
    if call.get("agent_id") != str(user["_id"]) and not is_admin_or_super(user):
        raise HTTPException(status_code=403, detail="You can only attach recordings to your own calls")
    content_type = (file.content_type or "").split(";")[0].strip().lower()
    ext = RECORDING_MIMES.get(content_type)
    if not ext:
        raise HTTPException(status_code=400, detail=f"Unsupported audio format: {content_type or 'unknown'}")
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Empty recording file")
    if len(data) > RECORDING_MAX_BYTES:
        raise HTTPException(status_code=413, detail="Recording too large (max 30 MB)")
    path = f"{STORAGE_APP_PREFIX}/recordings/{call['agent_id']}/{uuid.uuid4()}.{ext}"
    try:
        result = await asyncio.to_thread(_put_object, path, data, content_type)
    except Exception as e:
        logger.error(f"Recording upload failed: {e}")
        raise HTTPException(status_code=502, detail="Could not store recording. Please try again.")
    await db.calls.update_one(
        {"_id": ObjectId(call_id)},
        {"$set": {"has_recording": True, "recording_path": result["path"],
                  "recording_mime": content_type, "recording_size": len(data),
                  "recording_uploaded_at": now_iso()}})
    await db.activities.insert_one({
        "lead_id": call["lead_id"], "type": "recording", "message": "Call recording saved",
        "actor_id": str(user["_id"]), "actor_name": user.get("name"),
        "meta": {"call_id": call_id, "size": len(data)}, "created_at": now_iso()})
    return {"ok": True, "size": len(data)}


@api.get("/calls/{call_id}/recording")
async def get_recording(call_id: str, user: dict = Depends(get_current_user)):
    call = await db.calls.find_one({"_id": ObjectId(call_id)})
    if not call:
        raise HTTPException(status_code=404, detail="Call not found")
    if not _can_access_call(user, call):
        raise HTTPException(status_code=403, detail="You can only listen to your own call recordings")
    if not call.get("has_recording") or not call.get("recording_path"):
        raise HTTPException(status_code=404, detail="No recording for this call")
    try:
        data, content_type = await asyncio.to_thread(_get_object, call["recording_path"])
    except Exception as e:
        logger.error(f"Recording fetch failed: {e}")
        raise HTTPException(status_code=502, detail="Could not load recording")
    return Response(content=data, media_type=call.get("recording_mime") or content_type,
                    headers={"Cache-Control": "private, max-age=3600"})


# ---------------- Follow-ups ----------------
REMINDER_LEAD_MINUTES = 10


@api.get("/followups", response_model=List[Lead])
async def list_followups(user: dict = Depends(get_current_user),
                         agent_id: Optional[str] = None,
                         follow_up_status: Optional[str] = None,
                         date_from: Optional[str] = None,
                         date_to: Optional[str] = None):
    base_q = _lead_visibility_query(user)
    query = await _apply_visibility(base_q, user)
    query["follow_up_at"] = {"$ne": None}
    if (can_view_all(user) or user.get("role") == ROLE_TL) and agent_id:
        query["assigned_to"] = agent_id
    if follow_up_status:
        query["follow_up_status"] = follow_up_status
    if date_from or date_to:
        rng: dict = {"$ne": None}
        if date_from: rng["$gte"] = date_from
        if date_to: rng["$lte"] = date_to
        query["follow_up_at"] = rng
    docs = await db.leads.find(query).sort("follow_up_at", 1).to_list(1000)
    return [Lead.from_mongo(d) for d in docs]


@api.get("/followups/due", response_model=List[Lead])
async def due_followups(user: dict = Depends(get_current_user)):
    horizon = (datetime.now(timezone.utc) + timedelta(minutes=REMINDER_LEAD_MINUTES)).isoformat()
    base_q = _lead_visibility_query(user)
    query = await _apply_visibility(base_q, user)
    query["follow_up_at"] = {"$ne": None, "$lte": horizon}
    query["reminder_acked"] = {"$ne": True}
    docs = await db.leads.find(query).sort("follow_up_at", 1).to_list(100)
    return [Lead.from_mongo(d) for d in docs]


@api.post("/followups/{lead_id}/ack")
async def ack_followup(lead_id: str, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"_id": ObjectId(lead_id)})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    await _ensure_lead_access(lead, user)
    await db.leads.update_one({"_id": ObjectId(lead_id)}, {"$set": {"reminder_acked": True}})
    return {"ok": True}


@api.post("/followups/{lead_id}/complete")
async def complete_followup(lead_id: str, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"_id": ObjectId(lead_id)})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    await _ensure_lead_access(lead, user)
    if not lead.get("follow_up_at"):
        raise HTTPException(status_code=400, detail="No follow-up scheduled on this lead")
    await db.leads.update_one(
        {"_id": ObjectId(lead_id)},
        {"$set": {"follow_up_at": None, "follow_up_note": None, "reminder_acked": False,
                  "updated_at": now_iso()}})
    await _log_activity(lead_id, user, "followup", "Follow-up completed")
    return {"ok": True}


# ---------------- Settings ----------------
DEFAULT_SETTINGS = {
    "office_lat": float(os.environ.get("OFFICE_LAT", "28.4595")),
    "office_lng": float(os.environ.get("OFFICE_LNG", "77.0266")),
    "office_radius_m": int(os.environ.get("OFFICE_RADIUS_M", "500")),
    "office_start": os.environ.get("OFFICE_START", "11:00"),
    "office_end": os.environ.get("OFFICE_END", "18:00"),
    "office_label": os.environ.get("OFFICE_LABEL", "HQ"),
}


async def _get_settings() -> dict:
    doc = await db.settings.find_one({"_id": "office"})
    if not doc:
        doc = {"_id": "office", **DEFAULT_SETTINGS}
        await db.settings.insert_one(doc)
    return {k: doc.get(k, DEFAULT_SETTINGS[k]) for k in DEFAULT_SETTINGS}


@api.get("/settings")
async def get_settings(user: dict = Depends(get_current_user)):
    settings = await _get_settings()
    reset_doc = await db.settings.find_one({"_id": "operational_reset"})
    settings["data_reset_at"] = reset_doc.get("reset_at") if reset_doc else None
    settings["data_reset_by"] = reset_doc.get("reset_by") if reset_doc else None
    return settings


@api.put("/settings")
async def update_settings(payload: SettingsUpdate, actor: dict = Depends(require_super)):
    updates = payload.model_dump(exclude_none=True)
    if not updates:
        return await _get_settings()
    await db.settings.update_one({"_id": "office"}, {"$set": updates}, upsert=True)
    return await _get_settings()


@api.post("/settings/reset-data")
async def reset_operational_data(payload: SystemResetRequest, actor: dict = Depends(require_superadmin_only)):
    """
    Irreversibly wipes all call logs and attendance records so the system starts
    fresh from this moment. Requires the confirmation phrase 'RESET' (exact, case-sensitive).
    Leads, users, and office settings are left untouched — only call/attendance history
    and each lead's derived talk-time/call-count are cleared.
    """
    if (payload.confirm or "").strip() != "RESET":
        raise HTTPException(status_code=400, detail="Type RESET exactly to confirm this action.")

    calls_result = await db.calls.delete_many({})
    attendance_result = await db.attendance.delete_many({})
    await db.leads.update_many({}, {"$set": {"total_talk_time": 0, "call_count": 0}})

    reset_at = now_iso()
    await db.settings.update_one(
        {"_id": "operational_reset"},
        {"$set": {"reset_at": reset_at, "reset_by": actor.get("name")}},
        upsert=True,
    )
    return {
        "ok": True,
        "reset_at": reset_at,
        "calls_deleted": calls_result.deleted_count,
        "attendance_deleted": attendance_result.deleted_count,
    }


# ---------------- Attendance ----------------
def _parse_hhmm(s: str) -> Optional[tuple]:
    try:
        h, m = s.split(":")
        return int(h), int(m)
    except (ValueError, AttributeError):
        return None


def _expected_times_for(user: dict, day: _date, settings: dict) -> tuple:
    start = user.get("office_start") or settings["office_start"]
    end = user.get("office_end") or settings["office_end"]
    sh, sm = _parse_hhmm(start) or (11, 0)
    eh, em = _parse_hhmm(end) or (18, 0)
    # IST datetime for the given day
    ist = timezone(timedelta(minutes=330))
    exp_start = datetime(day.year, day.month, day.day, sh, sm, tzinfo=ist)
    exp_end = datetime(day.year, day.month, day.day, eh, em, tzinfo=ist)
    return exp_start.astimezone(timezone.utc), exp_end.astimezone(timezone.utc)


def _iso_to_dt(iso: str) -> datetime:
    return datetime.fromisoformat(iso.replace("Z", "+00:00"))


async def _upsert_attendance_calc(user_id: str) -> None:
    """Recompute overtime/late seconds for the row (used after edits)."""
    pass  # inline in checkout/edit


@api.post("/attendance/check-in")
async def attendance_check_in(payload: AttendanceCheckIn, user: dict = Depends(get_current_user)):
    if user.get("attendance_exempt"):
        raise HTTPException(status_code=400, detail="Your role does not require attendance marking")
    settings = await _get_settings()
    is_wfh = bool(user.get("wfh"))
    dist = haversine_m(payload.lat, payload.lng, settings["office_lat"], settings["office_lng"])
    if not is_wfh and dist > settings["office_radius_m"]:
        raise HTTPException(status_code=403,
                            detail=f"Outside office radius ({int(dist)} m > {settings['office_radius_m']} m). Please move closer to the office to mark attendance.")
    today = ist_today_str()
    existing = await db.attendance.find_one({"user_id": str(user["_id"]), "date": today})
    if existing and existing.get("check_in_at"):
        return {"ok": True, "already": True, "attendance_id": str(existing["_id"]),
                "distance_m": int(dist)}
    ist_day = ist_now().date()
    exp_start, exp_end = _expected_times_for(user, ist_day, settings)
    now = datetime.now(timezone.utc)
    late = max(0, int((now - exp_start).total_seconds()))
    doc = {
        "user_id": str(user["_id"]),
        "user_name": user.get("name"),
        "user_role": user.get("role"),
        "team_lead_id": user.get("team_lead_id"),
        "date": today,
        "check_in_at": now.isoformat(),
        "check_in_lat": payload.lat, "check_in_lng": payload.lng,
        "check_in_distance_m": int(dist),
        "check_in_accuracy_m": payload.accuracy,
        "check_in_wfh": is_wfh,
        "expected_start_at": exp_start.isoformat(),
        "expected_end_at": exp_end.isoformat(),
        "late_seconds": late,
        "status": "present",
        "created_at": now.isoformat(),
    }
    if existing:
        await db.attendance.update_one({"_id": existing["_id"]}, {"$set": doc})
        aid = str(existing["_id"])
    else:
        res = await db.attendance.insert_one(doc)
        aid = str(res.inserted_id)
    return {"ok": True, "attendance_id": aid, "distance_m": int(dist), "late_seconds": late}


@api.post("/attendance/check-out")
async def attendance_check_out(payload: AttendanceCheckOut, user: dict = Depends(get_current_user)):
    if user.get("attendance_exempt"):
        raise HTTPException(status_code=400, detail="Your role does not require attendance marking")
    settings = await _get_settings()
    is_wfh = bool(user.get("wfh"))
    dist = haversine_m(payload.lat, payload.lng, settings["office_lat"], settings["office_lng"])
    if not is_wfh and dist > settings["office_radius_m"]:
        raise HTTPException(status_code=403,
                            detail=f"Outside office radius ({int(dist)} m). Please check out from within the office.")
    today = ist_today_str()
    row = await db.attendance.find_one({"user_id": str(user["_id"]), "date": today})
    if not row or not row.get("check_in_at"):
        raise HTTPException(status_code=400, detail="You haven't checked in yet today")
    now = datetime.now(timezone.utc)
    checked_in = _iso_to_dt(row["check_in_at"])
    worked = max(0, int((now - checked_in).total_seconds()))
    exp_end = _iso_to_dt(row["expected_end_at"])
    overtime = max(0, int((now - exp_end).total_seconds()))
    await db.attendance.update_one(
        {"_id": row["_id"]},
        {"$set": {
            "check_out_at": now.isoformat(),
            "check_out_lat": payload.lat, "check_out_lng": payload.lng,
            "check_out_distance_m": int(dist),
            "worked_seconds": worked,
            "overtime_seconds": overtime,
        }})
    return {"ok": True, "worked_seconds": worked, "overtime_seconds": overtime}


@api.get("/attendance/me")
async def attendance_me(user: dict = Depends(get_current_user), days: int = 30):
    today = ist_now().date()
    start = (today - timedelta(days=days - 1)).isoformat()
    rows = await db.attendance.find(
        {"user_id": str(user["_id"]), "date": {"$gte": start}}
    ).sort("date", -1).to_list(200)
    settings = await _get_settings()
    for r in rows: r["_id"] = str(r["_id"])
    return {"records": rows, "settings": settings, "today": today.isoformat()}


@api.get("/attendance/today")
async def attendance_today(user: dict = Depends(get_current_user)):
    """List today's attendance (admins see everyone, TL sees their team). Filters attendance_exempt users."""
    today = ist_today_str()
    if can_view_all(user):
        users = await db.users.find({"active": True, "attendance_exempt": {"$ne": True}}).to_list(2000)
    elif user.get("role") == ROLE_TL:
        ids = await _team_member_ids(user)
        users = await db.users.find({"_id": {"$in": [ObjectId(i) for i in ids]}, "active": True,
                                     "attendance_exempt": {"$ne": True}}).to_list(500)
    else:
        raise HTTPException(status_code=403, detail="Not authorized")

    rows = await db.attendance.find({"date": today}).to_list(2000)
    row_by_uid = {r["user_id"]: r for r in rows}
    result = []
    for u in users:
        r = row_by_uid.get(str(u["_id"]))
        result.append({
            "user_id": str(u["_id"]),
            "user_name": u.get("name"),
            "username": u.get("username"),
            "role": u.get("role"),
            "team_lead_name": u.get("team_lead_name"),
            "wfh": bool(u.get("wfh")),
            "avatar_url": u.get("avatar_url"),
            "check_in_at": r.get("check_in_at") if r else None,
            "check_out_at": r.get("check_out_at") if r else None,
            "worked_seconds": r.get("worked_seconds", 0) if r else 0,
            "overtime_seconds": r.get("overtime_seconds", 0) if r else 0,
            "late_seconds": r.get("late_seconds", 0) if r else 0,
            "check_in_wfh": r.get("check_in_wfh", False) if r else False,
            "status": (r.get("status") if r else None) or ("absent" if not r else "present"),
            "attendance_id": str(r["_id"]) if r else None,
        })
    return {"date": today, "rows": result}


@api.get("/attendance/user/{user_id}")
async def attendance_user(user_id: str, days: int = 30, actor: dict = Depends(get_current_user)):
    target = await db.users.find_one({"_id": ObjectId(user_id)})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    if not can_view_all(actor):
        if actor.get("role") == ROLE_TL and target.get("team_lead_id") == str(actor["_id"]):
            pass
        elif str(target["_id"]) == str(actor["_id"]):
            pass
        else:
            raise HTTPException(status_code=403, detail="Not authorized")
    today = ist_now().date()
    start = (today - timedelta(days=days - 1)).isoformat()
    rows = await db.attendance.find(
        {"user_id": str(target["_id"]), "date": {"$gte": start}}
    ).sort("date", -1).to_list(500)
    for r in rows: r["_id"] = str(r["_id"])
    return {"user": UserPublic.from_mongo(target), "records": rows}


@api.put("/attendance/{att_id}")
async def edit_attendance(att_id: str, payload: AttendanceEdit, actor: dict = Depends(require_super)):
    row = await db.attendance.find_one({"_id": ObjectId(att_id)})
    if not row:
        raise HTTPException(status_code=404, detail="Attendance record not found")
    updates: dict = {}
    if payload.check_in_at is not None:
        updates["check_in_at"] = payload.check_in_at
    if payload.check_out_at is not None:
        updates["check_out_at"] = payload.check_out_at
    if payload.status:
        updates["status"] = payload.status
    if payload.note:
        updates["edit_note"] = payload.note
    updates["edited_by"] = actor.get("name")
    updates["edited_at"] = now_iso()
    # recompute worked/overtime if both times present
    ci = updates.get("check_in_at", row.get("check_in_at"))
    co = updates.get("check_out_at", row.get("check_out_at"))
    if ci and co:
        w = max(0, int((_iso_to_dt(co) - _iso_to_dt(ci)).total_seconds()))
        updates["worked_seconds"] = w
        exp_end = _iso_to_dt(row["expected_end_at"]) if row.get("expected_end_at") else None
        if exp_end:
            updates["overtime_seconds"] = max(0, int((_iso_to_dt(co) - exp_end).total_seconds()))
    await db.attendance.update_one({"_id": ObjectId(att_id)}, {"$set": updates})
    doc = await db.attendance.find_one({"_id": ObjectId(att_id)})
    doc["_id"] = str(doc["_id"])
    return doc


@api.post("/attendance/mark-manual")
async def mark_attendance_manual(user_id: str = Form(...), date_str: str = Form(...),
                                 status: str = Form("present"), note: Optional[str] = Form(None),
                                 actor: dict = Depends(require_super)):
    """Superadmin creates/updates an attendance record for someone."""
    target = await db.users.find_one({"_id": ObjectId(user_id)})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    existing = await db.attendance.find_one({"user_id": user_id, "date": date_str})
    payload = {
        "user_id": user_id, "user_name": target.get("name"),
        "user_role": target.get("role"),
        "team_lead_id": target.get("team_lead_id"),
        "date": date_str, "status": status,
        "edit_note": note or f"Manually marked {status} by {actor.get('name')}",
        "edited_by": actor.get("name"), "edited_at": now_iso(),
    }
    if existing:
        await db.attendance.update_one({"_id": existing["_id"]}, {"$set": payload})
        aid = str(existing["_id"])
    else:
        payload["created_at"] = now_iso()
        res = await db.attendance.insert_one(payload)
        aid = str(res.inserted_id)
    return {"ok": True, "id": aid}


@api.get("/attendance/stats")
async def attendance_stats(actor: dict = Depends(get_current_user), period: str = "week"):
    """period: 'week' | 'month'. Returns absentees, top overtime, top late."""
    if not (can_view_all(actor) or actor.get("role") == ROLE_TL):
        raise HTTPException(status_code=403, detail="Not authorized")
    today = ist_now().date()
    if period == "month":
        start = today.replace(day=1)
    else:
        start = today - timedelta(days=today.weekday())  # monday
    start_str = start.isoformat()

    # If an operational data reset happened inside this period, don't count days
    # before the reset as "absent" — those days' real records were wiped, not missed.
    reset_doc = await db.settings.find_one({"_id": "operational_reset"})
    effective_start = start
    if reset_doc and reset_doc.get("reset_at"):
        reset_date = (_iso_to_dt(reset_doc["reset_at"]) + timedelta(minutes=330)).date()
        if reset_date > effective_start:
            effective_start = reset_date

    # Users in scope
    if can_view_all(actor):
        users = await db.users.find({"active": True, "attendance_exempt": {"$ne": True},
                                     "role": {"$in": list(ATTENDANCE_REQUIRED_ROLES)}}).to_list(2000)
    else:
        ids = await _team_member_ids(actor)
        users = await db.users.find({"_id": {"$in": [ObjectId(i) for i in ids]},
                                     "active": True, "attendance_exempt": {"$ne": True},
                                     "role": {"$in": list(ATTENDANCE_REQUIRED_ROLES)}}).to_list(500)

    ids = [str(u["_id"]) for u in users]
    rows = await db.attendance.find(
        {"user_id": {"$in": ids}, "date": {"$gte": start_str}}
    ).to_list(5000)

    from collections import defaultdict
    by_user = defaultdict(list)
    for r in rows:
        by_user[r["user_id"]].append(r)

    working_days_count = 0
    d = effective_start
    while d <= today:
        working_days_count += 1
        d += timedelta(days=1)

    per_user = []
    for u in users:
        uid = str(u["_id"])
        user_rows = by_user.get(uid, [])
        present = sum(1 for r in user_rows if r.get("status") == "present")
        # Only count days on/after this person's own working-day baseline
        # (joining date, if it's later than the period/reset floor) as expected days.
        person_floor = effective_start
        joining = u.get("joining_date")
        if joining:
            try:
                jd = datetime.fromisoformat(joining).date()
                if jd > person_floor:
                    person_floor = jd
            except (ValueError, TypeError):
                pass
        allowed_days = set(u.get("working_days") or [0, 1, 2, 3, 4, 5])
        expected = 0
        d = person_floor
        while d <= today:
            if d.weekday() in allowed_days:
                expected += 1
            d += timedelta(days=1)
        absent = max(0, expected - present)
        overtime = sum(r.get("overtime_seconds", 0) for r in user_rows)
        late = sum(r.get("late_seconds", 0) for r in user_rows)
        late_days = sum(1 for r in user_rows if (r.get("late_seconds", 0) or 0) > 300)  # >5 min
        worked = sum(r.get("worked_seconds", 0) for r in user_rows)
        per_user.append({
            "user_id": uid, "name": u.get("name"), "username": u.get("username"),
            "role": u.get("role"), "team_lead_name": u.get("team_lead_name"),
            "present_days": present, "absent_days": absent, "late_days": late_days,
            "overtime_seconds": overtime, "late_seconds": late, "worked_seconds": worked,
        })

    top_overtime = sorted(per_user, key=lambda x: -x["overtime_seconds"])[:5]
    top_absent = sorted([x for x in per_user if x["absent_days"] > 0],
                        key=lambda x: -x["absent_days"])[:10]
    top_late = sorted(per_user, key=lambda x: -x["late_seconds"])[:5]
    on_time = [x for x in per_user if x["late_days"] == 0 and x["present_days"] > 0]
    return {
        "period": period, "start": start_str, "end": today.isoformat(),
        "working_days": working_days_count,
        "per_user": per_user,
        "top_overtime": top_overtime,
        "top_absent": top_absent,
        "top_late": top_late,
        "on_time_count": len(on_time),
    }


# ---------------- Team chat / groups ----------------
@api.get("/groups")
async def list_groups(user: dict = Depends(get_current_user)):
    docs = await db.groups.find({"member_ids": str(user["_id"])}).sort("created_at", -1).to_list(200)
    for d in docs:
        d["_id"] = str(d["_id"])
        last = await db.messages.find_one({"group_id": str(d["_id"])}, sort=[("created_at", -1)])
        d["last_message"] = last.get("text") if last else None
        d["last_at"] = last.get("created_at") if last else d.get("created_at")
    return docs


@api.post("/groups")
async def create_group(payload: GroupCreate, user: dict = Depends(get_current_user)):
    member_ids = list(set((payload.member_ids or []) + [str(user["_id"])]))
    users = await db.users.find({"_id": {"$in": [ObjectId(i) for i in member_ids]}}).to_list(200)
    doc = {
        "name": payload.name, "member_ids": [str(u["_id"]) for u in users],
        "member_names": [u.get("name") for u in users],
        "created_by": str(user["_id"]), "created_by_name": user.get("name"),
        "created_at": now_iso(),
    }
    res = await db.groups.insert_one(doc)
    doc["_id"] = str(res.inserted_id)
    return doc


@api.put("/groups/{group_id}")
async def update_group(group_id: str, payload: GroupUpdate, user: dict = Depends(get_current_user)):
    g = await db.groups.find_one({"_id": ObjectId(group_id)})
    if not g:
        raise HTTPException(status_code=404, detail="Group not found")
    if str(user["_id"]) not in g.get("member_ids", []) and not is_admin_or_super(user):
        raise HTTPException(status_code=403, detail="Not a member of this group")
    updates: dict = {}
    if payload.name: updates["name"] = payload.name
    if payload.member_ids is not None:
        users = await db.users.find({"_id": {"$in": [ObjectId(i) for i in payload.member_ids]}}).to_list(200)
        updates["member_ids"] = [str(u["_id"]) for u in users]
        updates["member_names"] = [u.get("name") for u in users]
    if updates:
        await db.groups.update_one({"_id": ObjectId(group_id)}, {"$set": updates})
    doc = await db.groups.find_one({"_id": ObjectId(group_id)})
    doc["_id"] = str(doc["_id"])
    return doc


@api.get("/groups/{group_id}/messages")
async def list_messages(group_id: str, user: dict = Depends(get_current_user), limit: int = 200):
    g = await db.groups.find_one({"_id": ObjectId(group_id)})
    if not g:
        raise HTTPException(status_code=404, detail="Group not found")
    if str(user["_id"]) not in g.get("member_ids", []) and not is_admin_or_super(user):
        raise HTTPException(status_code=403, detail="Not a member of this group")
    docs = await db.messages.find({"group_id": group_id}).sort("created_at", 1).to_list(limit)
    for d in docs: d["_id"] = str(d["_id"])
    return docs


@api.post("/groups/{group_id}/messages")
async def post_message(group_id: str, payload: MessageCreate, user: dict = Depends(get_current_user)):
    g = await db.groups.find_one({"_id": ObjectId(group_id)})
    if not g:
        raise HTTPException(status_code=404, detail="Group not found")
    if str(user["_id"]) not in g.get("member_ids", []) and not is_admin_or_super(user):
        raise HTTPException(status_code=403, detail="Not a member of this group")
    doc = {
        "group_id": group_id, "sender_id": str(user["_id"]),
        "sender_name": user.get("name"), "sender_role": user.get("role"),
        "text": payload.text, "created_at": now_iso(),
    }
    res = await db.messages.insert_one(doc)
    doc["_id"] = str(res.inserted_id)
    return doc


# ---------------- Alerts (broadcast) ----------------
@api.get("/alerts")
async def list_alerts(user: dict = Depends(get_current_user), limit: int = 50):
    role = user.get("role")
    uid = str(user["_id"])
    now = now_iso()
    docs = await db.alerts.find(
        {"$or": [{"target": "all"}, {"target": role}, {"target": uid}],
         "$and": [{"$or": [{"expires_at": None}, {"expires_at": {"$gt": now}}]}]}
    ).sort("created_at", -1).to_list(limit)
    for d in docs: d["_id"] = str(d["_id"])
    return docs


@api.post("/alerts")
async def create_alert(payload: AlertCreate, user: dict = Depends(get_current_user)):
    duration = max(1, int(payload.duration_hours or 24))
    expires = (datetime.now(timezone.utc) + timedelta(hours=duration)).isoformat()
    doc = {
        "message": payload.message, "target": payload.target, "priority": payload.priority,
        "duration_hours": duration, "expires_at": expires,
        "from_id": str(user["_id"]), "from_name": user.get("name"),
        "from_role": user.get("role"), "created_at": now_iso(),
    }
    res = await db.alerts.insert_one(doc)
    doc["_id"] = str(res.inserted_id)
    return doc


# ---------------- Dashboards ----------------
async def _agent_stats(agent_id: str) -> dict:
    leads = await db.leads.find({"assigned_to": agent_id}).to_list(5000)
    calls = await db.calls.find({"agent_id": agent_id, "status": "completed"}).to_list(5000)
    won = len([l for l in leads if l.get("status") == "won"])
    talk = sum(c.get("duration", 0) for c in calls)
    return {
        "leads": len(leads),
        "won": won,
        "lost": len([l for l in leads if l.get("status") == "lost"]),
        "active": len([l for l in leads if l.get("status") not in ("won", "lost")]),
        "calls": len(calls),
        "talk_time": talk,
        "avg_call": int(talk / len(calls)) if calls else 0,
        "conversion": round(won / len(leads) * 100, 1) if leads else 0.0,
        "pipeline_value": sum(l.get("budget") or 0 for l in leads if l.get("status") not in ("won", "lost")),
        "won_value": sum(l.get("budget") or 0 for l in leads if l.get("status") == "won"),
    }


@api.get("/dashboard/admin")
async def admin_dashboard(user: dict = Depends(get_current_user)):
    # Team leads see their own team's admin-style dashboard
    is_full = can_view_all(user)
    if not is_full and user.get("role") != ROLE_TL:
        raise HTTPException(status_code=403, detail="Not authorized")

    if is_full:
        leads = await db.leads.find({}).to_list(20000)
        calls = await db.calls.find({"status": "completed"}).to_list(20000)
        agents = await db.users.find({"role": {"$in": [ROLE_SALES, ROLE_TL]}}).to_list(2000)
    else:
        ids = await _team_member_ids(user)
        leads = await db.leads.find({"assigned_to": {"$in": ids}}).to_list(20000)
        calls = await db.calls.find({"agent_id": {"$in": ids}, "status": "completed"}).to_list(20000)
        agents = await db.users.find({"_id": {"$in": [ObjectId(i) for i in ids]},
                                      "role": {"$in": [ROLE_SALES, ROLE_TL]}}).to_list(500)

    by_status = {s: 0 for s in LEAD_STATUSES}
    by_source: dict = {}
    for l in leads:
        by_status[l.get("status", "new")] = by_status.get(l.get("status", "new"), 0) + 1
        by_source[l.get("source") or "Unknown"] = by_source.get(l.get("source") or "Unknown", 0) + 1

    total_talk = sum(c.get("duration", 0) for c in calls)
    won = by_status.get("won", 0)

    leaderboard = []
    for a in agents:
        s = await _agent_stats(str(a["_id"]))
        leaderboard.append({"agent_id": str(a["_id"]), "name": a["name"], "username": a["username"],
                            "role": a.get("role"), "team_lead_name": a.get("team_lead_name"),
                            "active": a.get("active", True), **s})
    leaderboard.sort(key=lambda x: (-x["won"], -x["talk_time"]))

    trend = []
    today = datetime.now(timezone.utc).date()
    for i in range(13, -1, -1):
        day = (today - timedelta(days=i)).isoformat()
        day_leads = len([l for l in leads if (l.get("created_at") or "").startswith(day)])
        day_talk = sum(c.get("duration", 0) for c in calls if (c.get("started_at") or "").startswith(day))
        day_calls = len([c for c in calls if (c.get("started_at") or "").startswith(day)])
        trend.append({"date": day[5:], "leads": day_leads, "talk_minutes": round(day_talk / 60, 1),
                      "calls": day_calls})

    return {
        "scope": "full" if is_full else "team",
        "kpis": {
            "total_leads": len(leads),
            "unassigned": len([l for l in leads if not l.get("assigned_to")]),
            "active_leads": len([l for l in leads if l.get("status") not in ("won", "lost")]),
            "won": won,
            "conversion": round(won / len(leads) * 100, 1) if leads else 0.0,
            "total_calls": len(calls),
            "total_talk_time": total_talk,
            "avg_call": int(total_talk / len(calls)) if calls else 0,
            "agents": len(agents),
            "active_agents": len([a for a in agents if a.get("active", True)]),
            "pipeline_value": sum(l.get("budget") or 0 for l in leads if l.get("status") not in ("won", "lost")),
            "won_value": sum(l.get("budget") or 0 for l in leads if l.get("status") == "won"),
        },
        "by_status": [{"status": k, "count": v} for k, v in by_status.items()],
        "by_source": [{"source": k, "count": v} for k, v in by_source.items()],
        "leaderboard": leaderboard,
        "trend": trend,
        "recent_calls": [{**c, "_id": str(c["_id"])} for c in
                         sorted(calls, key=lambda x: x.get("started_at", ""), reverse=True)[:8]],
    }


@api.get("/dashboard/agent")
async def agent_dashboard(user: dict = Depends(get_current_user)):
    agent_id = str(user["_id"])
    stats = await _agent_stats(agent_id)
    leads = await db.leads.find({"assigned_to": agent_id}).sort("updated_at", -1).to_list(2000)
    calls = await db.calls.find({"agent_id": agent_id, "status": "completed"}).to_list(2000)
    by_status = {s: 0 for s in LEAD_STATUSES}
    for l in leads:
        by_status[l.get("status", "new")] = by_status.get(l.get("status", "new"), 0) + 1
    trend = []
    today = datetime.now(timezone.utc).date()
    for i in range(13, -1, -1):
        day = (today - timedelta(days=i)).isoformat()
        trend.append({
            "date": day[5:],
            "calls": len([c for c in calls if (c.get("started_at") or "").startswith(day)]),
            "talk_minutes": round(sum(c.get("duration", 0) for c in calls
                                      if (c.get("started_at") or "").startswith(day)) / 60, 1),
        })
    return {
        "stats": stats,
        "by_status": [{"status": k, "count": v} for k, v in by_status.items()],
        "trend": trend,
        "my_leads": [Lead.from_mongo(l) for l in leads[:10]],
    }


@api.get("/dashboard/agent/{agent_id}")
async def agent_detail(agent_id: str, actor: dict = Depends(get_current_user)):
    agent = await db.users.find_one({"_id": ObjectId(agent_id)})
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")
    if not can_view_all(actor):
        if actor.get("role") == ROLE_TL and agent.get("team_lead_id") == str(actor["_id"]):
            pass
        elif str(agent["_id"]) == str(actor["_id"]):
            pass
        else:
            raise HTTPException(status_code=403, detail="Not authorized")
    stats = await _agent_stats(agent_id)
    leads = await db.leads.find({"assigned_to": agent_id}).sort("updated_at", -1).to_list(1000)
    calls = await db.calls.find({"agent_id": agent_id}).sort("started_at", -1).to_list(200)
    for c in calls: c["_id"] = str(c["_id"])
    return {"agent": UserPublic.from_mongo(agent), "stats": stats,
            "leads": [Lead.from_mongo(l) for l in leads], "calls": calls}


# ---------------- Daily report ----------------
def _fmt_dur(sec) -> str:
    sec = int(sec or 0)
    h, m = sec // 3600, (sec % 3600) // 60
    if h: return f"{h}h {m}m"
    if m: return f"{m}m {sec % 60}s"
    return f"{sec}s"


# --- Reusable daily report builder (formerly the /reports/daily route) ---
async def _build_daily_report(tz_offset: int = 0) -> dict:
    utc_now = datetime.now(timezone.utc)
    local_now = utc_now - timedelta(minutes=tz_offset)
    local_midnight = local_now.replace(hour=0, minute=0, second=0, microsecond=0)
    today_start = local_midnight + timedelta(minutes=tz_offset)
    yesterday_start = today_start - timedelta(days=1)
    today_end = today_start + timedelta(days=1)
    y0, t0, t1 = yesterday_start.isoformat(), today_start.isoformat(), today_end.isoformat()

    calls = await db.calls.find({"status": "completed",
                                 "started_at": {"$gte": y0, "$lt": t0}}).to_list(5000)
    per_agent: dict = {}
    for c in calls:
        a = per_agent.setdefault(c.get("agent_name") or "Unknown", {"calls": 0, "talk_time": 0})
        a["calls"] += 1
        a["talk_time"] += c.get("duration", 0)
    per_agent_list = [{"name": k, **v} for k, v in
                      sorted(per_agent.items(), key=lambda x: -x[1]["talk_time"])]
    total_talk = sum(c.get("duration", 0) for c in calls)
    new_leads = await db.leads.count_documents({"created_at": {"$gte": y0, "$lt": t0}})

    overdue_docs = await db.leads.find(
        {"follow_up_at": {"$ne": None, "$lt": utc_now.isoformat()}}).sort("follow_up_at", 1).to_list(500)
    today_docs = await db.leads.find(
        {"follow_up_at": {"$gte": utc_now.isoformat(), "$lt": t1}}).sort("follow_up_at", 1).to_list(500)

    def loc_time(iso):
        try:
            d = datetime.fromisoformat(iso)
            return (d - timedelta(minutes=tz_offset)).strftime("%I:%M %p").lstrip("0").lower()
        except (ValueError, TypeError):
            return iso or ""

    def fu_row(lead):
        return {"lead_id": str(lead["_id"]), "name": lead["name"], "phone": lead["phone"],
                "agent": lead.get("assigned_to_name") or "Unassigned",
                "at": lead.get("follow_up_at"), "at_local": loc_time(lead.get("follow_up_at")),
                "note": lead.get("follow_up_note")}

    today_fu = [fu_row(x) for x in today_docs]
    overdue_fu = [fu_row(x) for x in overdue_docs]

    date_label = local_now.strftime("%a, %d %b %Y")
    ydate_label = (local_now - timedelta(days=1)).strftime("%d %b")

    lines = [
        "🏢 *UNIQUE PRIME REALITY — Daily Report*",
        f"📅 {date_label}", "",
        f"📞 *Yesterday ({ydate_label}):* {len(calls)} calls · {_fmt_dur(total_talk)} talk time · {new_leads} new leads",
    ]
    if per_agent_list:
        lines += ["", "*Talk time per salesperson:*"]
        for a in per_agent_list:
            lines.append(f"• {a['name']} — {a['calls']} calls · {_fmt_dur(a['talk_time'])}")
    lines += ["", f"⏰ *Today's follow-ups ({len(today_fu)}):*"]
    if today_fu:
        for f in today_fu[:15]:
            lines.append(f"• {f['at_local']} — {f['name']} ({f['agent']})")
    else:
        lines.append("• None scheduled")
    if overdue_fu:
        lines += ["", f"🔴 *Overdue follow-ups ({len(overdue_fu)}):*"]
        for f in overdue_fu[:10]:
            lines.append(f"• {f['name']} ({f['agent']})")
    lines += ["", "— Unique Prime Reality CRM"]

    return {
        "date": date_label,
        "yesterday": {"calls": len(calls), "talk_time": total_talk,
                      "per_agent": per_agent_list, "new_leads": new_leads},
        "today_followups": today_fu,
        "overdue_followups": overdue_fu,
        "whatsapp_text": "\n".join(lines),
    }


# --- JSON daily report (keeps the original route) ---
@api.get("/reports/daily")
async def daily_report(admin: dict = Depends(require_admin), tz_offset: int = 0):
    return await _build_daily_report(tz_offset)


# --- PDF daily report (new route) ---
@api.get("/reports/daily/pdf")
async def daily_report_pdf(admin: dict = Depends(require_admin), tz_offset: int = 0):
    data = await _build_daily_report(tz_offset)

    # reports_pdf.py expects each agent's talk_time as a formatted string
    # ("46m 10s") and follow-up counts as plain ints — the JSON route above
    # returns raw seconds and full lists, so adapt the shape here.
    pdf_report = {
        "date": data["date"],
        "generated_for": (datetime.now(timezone.utc) - timedelta(minutes=tz_offset)).strftime("%a, %d %b %Y"),
        "yesterday": {
            "new_leads": data["yesterday"]["new_leads"],
            "per_agent": [
                {"name": a["name"], "calls": a["calls"], "talk_time": _fmt_dur(a["talk_time"])}
                for a in data["yesterday"]["per_agent"]
            ] or [{"name": "No calls yet", "calls": 0, "talk_time": "0s"}],
        },
        "today_followups": len(data["today_followups"]),
        "overdue_followups": len(data["overdue_followups"]),
    }

    pdf_bytes = generate_daily_report_pdf(pdf_report)
    safe_date = pdf_report["date"].replace(" ", "-").replace(",", "")
    filename = f"daily-report-{safe_date}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


@api.get("/")
async def root():
    return {"service": "Unique Prime Reality CRM", "status": "ok"}


app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---------------- Startup seed ----------------


@app.on_event("startup")
async def startup():
    await db.users.create_index("username", unique=True)
    await db.leads.create_index("assigned_to")
    await db.leads.create_index("status")
    await db.calls.create_index("agent_id")
    await db.activities.create_index("lead_id")
    await db.attendance.create_index([("user_id", 1), ("date", 1)], unique=True)
    await db.attendance.create_index("date")
    await db.messages.create_index([("group_id", 1), ("created_at", 1)])
    await db.notifications.create_index([("user_id", 1), ("created_at", -1)])

    # ONE-TIME PURGE — delete all demo/seed data on first boot of this release.
    purge_flag = await db.settings.find_one({"_id": "seed_purged_v2"})
    if not purge_flag:
        keep = {os.environ["ADMIN_EMAIL"].lower(), os.environ["SUPERADMIN_EMAIL"].lower()}
        await db.users.delete_many({"username": {"$nin": list(keep)}})
        await db.leads.delete_many({})
        await db.calls.delete_many({})
        await db.activities.delete_many({})
        await db.attendance.delete_many({})
        await db.messages.delete_many({})
        await db.groups.delete_many({})
        await db.alerts.delete_many({})
        await db.notifications.delete_many({})
        await db.settings.insert_one({"_id": "seed_purged_v2", "at": now_iso()})
        logger.info("Purged all demo/seed data — clean install")

    async def _seed(username, password, name, role, email, phone):
        username = username.strip().lower()
        existing = await db.users.find_one({"username": username})
        base = {
            "username": username, "name": name, "email": email, "phone": phone, "role": role,
            "office_start": os.environ.get("OFFICE_START", "11:00"),
            "office_end": os.environ.get("OFFICE_END", "18:00"),
            "working_days": [0, 1, 2, 3, 4, 5],
            "active": True,
        }
        if not existing:
            base["password_hash"] = hash_password(password)
            base["created_at"] = now_iso()
            await db.users.insert_one(base)
            logger.info(f"Seeded founder: {username} [{role}]")
        else:
            updates = {"name": name, "role": role}
            await db.users.update_one({"_id": existing["_id"]}, {"$set": updates})

    # Only two founder accounts are seeded. Everyone else is created from the Team page.
    await _seed(os.environ["SUPERADMIN_EMAIL"], os.environ["SUPERADMIN_PASSWORD"],
                os.environ["SUPERADMIN_NAME"], ROLE_SUPER,
                "vranda@uniqueprimereality.com", os.environ.get("SUPERADMIN_PHONE"))
    await _seed(os.environ["ADMIN_EMAIL"], os.environ["ADMIN_PASSWORD"],
                os.environ["ADMIN_NAME"], ROLE_ADMIN,
                "sandeep@uniqueprimereality.com", os.environ.get("ADMIN_PHONE"))

    await db.users.update_one({"username": os.environ["ADMIN_EMAIL"].lower()},
                              {"$set": {"attendance_exempt": True, "wfh": False,
                                        "joining_date": "2020-04-01"}})
    await db.users.update_one({"username": os.environ["SUPERADMIN_EMAIL"].lower()},
                              {"$set": {"attendance_exempt": False, "wfh": True,
                                        "joining_date": "2021-06-15"}})

    # Seed an "All Hands" group with just the two founders (grows automatically when
    # new profiles are added). Skipped if it already exists.
    if await db.groups.count_documents({}) == 0:
        founders = await db.users.find({"username": {"$in": list(
            {os.environ["ADMIN_EMAIL"].lower(), os.environ["SUPERADMIN_EMAIL"].lower()})}}).to_list(10)
        if founders:
            await db.groups.insert_one({
                "name": "All Hands",
                "member_ids": [str(u["_id"]) for u in founders],
                "member_names": [u.get("name") for u in founders],
                "created_by": str(founders[0]["_id"]),
                "created_by_name": founders[0].get("name"),
                "created_at": now_iso(),
            })


@app.on_event("shutdown")
async def shutdown():
    client.close()
